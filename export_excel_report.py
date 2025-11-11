"""Create a polished Excel report from cross-validated and single-run CSVs.

Produces: model_performance_report.xlsx with sheets:
- CV Summary: sorted CV table (Accuracy Mean desc), formatted and color-scaled
- Single Run: the single train/test run table

Also adds a bar chart of Accuracy (%) Mean.
"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference


def make_report(cv_csv='model_performance_comparison_cv.csv', single_csv='model_performance_comparison.csv', out='model_performance_report.xlsx'):
    p_cv = Path(cv_csv)
    p_single = Path(single_csv)
    if not p_cv.exists():
        raise FileNotFoundError(p_cv)
    if not p_single.exists():
        raise FileNotFoundError(p_single)

    df_cv = pd.read_csv(p_cv)
    # sort by Accuracy (%) Mean desc
    if 'Accuracy (%) Mean' in df_cv.columns:
        df_cv = df_cv.sort_values(by='Accuracy (%) Mean', ascending=False).reset_index(drop=True)

    df_single = pd.read_csv(p_single)

    # write to Excel with two sheets
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df_cv.to_excel(writer, sheet_name='CV Summary', index=False)
        df_single.to_excel(writer, sheet_name='Single Run', index=False)

    # open with openpyxl for formatting and chart
    wb = load_workbook(out)
    ws = wb['CV Summary']

    # header style
    header_font = Font(bold=True)
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # set column widths a bit
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Apply color scale rule to Accuracy (%) Mean column if present
    headers = [c.value for c in ws[1]]
    if 'Accuracy (%) Mean' in headers:
        idx = headers.index('Accuracy (%) Mean') + 1  # 1-based
        col_letter = ws.cell(row=1, column=idx).column_letter
        data_range = f"{col_letter}2:{col_letter}{ws.max_row}"
        rule = ColorScaleRule(start_type='min', start_color='FFF8CB', mid_type='percentile', mid_value=50, mid_color='FFD966', end_type='max', end_color='63BE7B')
        ws.conditional_formatting.add(data_range, rule)

    # Add a bar chart of Accuracy (%) Mean
    try:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Accuracy (%) Mean by Model"
        chart.y_axis.title = 'Accuracy (%)'
        chart.x_axis.title = 'Model'

        # Accuracy values
        acc_col = headers.index('Accuracy (%) Mean') + 1
        data = Reference(ws, min_col=acc_col, min_row=1, max_row=ws.max_row)
        # categories = model names
        cat_col = headers.index('Model') + 1
        cats = Reference(ws, min_col=cat_col, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws.add_chart(chart, f"{ws.cell(row=1, column=ws.max_column).column_letter}{ws.max_row + 2}")
    except Exception:
        # ignore charting errors
        pass

    wb.save(out)
    return Path(out)


if __name__ == '__main__':
    out = make_report()
    print('Saved report to', out)
